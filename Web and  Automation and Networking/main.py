from calendar import c
from types import CellType
import xlsxwriter
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
def driverClose():
    global driver
    driver.close()
def driverOpen():
    global driver
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--incognito")
    driver = webdriver.Chrome(executable_path="C:\chromedriver.exe",options=chrome_options)
    driver.maximize_window()
# get all data in sheet
file = "File.xlsx"
wb_obj = openpyxl.load_workbook(file)
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
outWorkbook = xlsxwriter.Workbook('data.xlsx')
outSheet = outWorkbook.add_worksheet()
outSheet.write("A1", "Company Name")
outSheet.write("B1", "Ceo name")
driverOpen()
temp=1
try:
    for i in range(2, m_row + 1):
            companyName = sheet_obj.cell(row=i, column=1).value
            print(i," / ",m_row," ",companyName)
            url="https://www.google.com/search?q="+companyName
            driver.get(url)
            try:
                CeoName=driver.find_element_by_xpath("//div[@role='heading']//div//a").text
                CeoName=CeoName.lower()
                CeoName=CeoName.replace("ceo","")
                CeoName=CeoName.replace(".com","")
                CeoName=CeoName.replace("-","")
                outSheet.write(i - 1, 0,companyName)
                outSheet.write(i - 1, 1, CeoName)
            except:
                try:
                    CeoName=driver.find_element_by_xpath("//div[@data-tts='answers']").text
                    CeoName=CeoName.lower()
                    CeoName=CeoName.replace("ceo","")
                    CeoName=CeoName.replace(".com","")
                    CeoName=CeoName.replace("-","")
                    outSheet.write(i - 1, 0,companyName)
                    outSheet.write(i - 1, 1, CeoName)
                except:
                    try:
                        CeoName=driver.find_element_by_xpath("//div[@role='heading']//span[@lang='en']//span//b").text
                        CeoName=CeoName.lower()
                        CeoName=CeoName.replace("ceo","")
                        CeoName=CeoName.replace(".com","")
                        CeoName=CeoName.replace("-","")
                        outSheet.write(i - 1, 0,companyName)
                        outSheet.write(i - 1, 1, CeoName)
                    except:
                        CeoName=""
                        outSheet.write(i - 1, 0,companyName)
                        outSheet.write(i - 1, 1, CeoName)
            print(CeoName)
            temp+=1
            if(temp>5):
                driverClose()
                temp=1
                driverOpen()
except:
    outWorkbook.close()
    print("Saving...")
    exit()
outWorkbook.close()